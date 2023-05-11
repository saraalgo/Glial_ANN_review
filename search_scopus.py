# To manage data
import pandas as pd
import numpy as np
import requests
from pyscopus import Scopus
# To give format to xls with searched papers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl import Workbook


# 1. Add API KEY
# ===
key = "3ad5683c6a6290935a960e4159176538"
scopus = Scopus(key)

# 2. Run queries
# ===

## Query 1: astrocyte neural networks + glia
search_df1 = scopus.search("TITLE-ABS-KEY(artificial astrocytes neural networks) AND TITLE-ABS-KEY(glia)", count=50, view='STANDARD')
search_df1[["subtype_description", "title", "publication_name", "doi", "cover_date", "scopus_id"]]

## Query 2: connectionist system + astrocyte (glial?)
search_df2 = scopus.search("TITLE-ABS-KEY(connectionist system) AND TITLE-ABS-KEY(astrocyte)", count=50, view='STANDARD')
search_df2[["subtype_description", "title", "publication_name", "doi", "cover_date", "scopus_id"]]

## Query 3: glial networks + perceptron
search_df3 = scopus.search("TITLE-ABS-KEY(glial networks) AND TITLE-ABS-KEY(perceptron)", count=50, view='STANDARD')
search_df3[["subtype_description", "title", "publication_name", "doi", "cover_date", "scopus_id"]]

## Query 4: neuro-glial + artificial network
search_df4 = scopus.search("TITLE-ABS-KEY(neuro-glial) AND TITLE-ABS-KEY(artificial network)", count=100, view='STANDARD')
search_df4[["subtype_description", "title", "publication_name", "doi", "cover_date", "scopus_id"]]

## Query 5: neuroglial + artificial network
search_df5 = scopus.search("TITLE-ABS-KEY(neuroglial) AND TITLE-ABS-KEY(artificial network)", count=100, view='STANDARD')
search_df5[["subtype_description", "title", "publication_name", "doi", "cover_date", "scopus_id"]]

## Query 6: neuron-glia + artificial network
search_df6 = scopus.search("TITLE-ABS-KEY(neuron-glia) AND TITLE-ABS-KEY(artificial network)", count=100, view='STANDARD')
search_df6[["subtype_description", "title", "publication_name", "doi", "cover_date", "scopus_id"]]

## Query 7: neuron-astrocyte + artificial network
search_df7 = scopus.search("TITLE-ABS-KEY(neuron-astrocyte) AND TITLE-ABS-KEY(artificial network)", count=100, view='STANDARD')
search_df7[["subtype_description", "title", "publication_name", "doi", "cover_date", "scopus_id"]]


# 3. Merge all queries and remove duplicated papers
# ===
search_df = pd.concat([search_df1, search_df2, search_df3, search_df4, search_df5, search_df6, search_df7], axis=0)
search_df = search_df.drop_duplicates(subset=['title'])

# 4. Save 
# ===
scopus_search = search_df[["subtype_description", "title", "publication_name", "doi", "cover_date", "scopus_id"]]

new_names = {'subtype_description': 'Publication Type', 'title': 'Title', 'publication_name': 'Source',
            'doi': 'DOI', 'cover_date': 'Date', 'scopus_id': 'Scopus ID'}
scopus_search = scopus_search.rename(columns=new_names)
scopus_search['Date'] = scopus_search['Date'].str.split('-').str[0]


wb = Workbook()
ws = wb.active

for r, row in enumerate(dataframe_to_rows(scopus_search, index=False, header=True)):
    for c, val in enumerate(row):
        # write the cell value
        ws.cell(row=r+1, column=c+1, value=val)
        
        # set bold font for the column headers
        if r == 0:
            ws.cell(row=r+1, column=c+1).font = Font(bold=True)

# set column width based on the maximum length of values in each column
for i, column in enumerate(scopus_search.columns):
    column_width = max(scopus_search[column].apply(lambda x: len(str(x)))) + 2
    ws.column_dimensions[chr(i+65)].width = column_width

# save the workbook to an Excel file
wb.save('scopus_search.xlsx')